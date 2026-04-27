#!/usr/bin/env python3
"""Génère le fichier Excel Business Plan BarCoins depuis les données CFO."""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

OUTPUT = '/mnt/c/Users/jean-/OneDrive/Documents/Barcoins/07_EXPORTS/xlsx/impact-economique-bcoins-avril-2026.xlsx'
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

# Couleurs BarCoins
NAVY   = '0D1B2A'
GOLD   = 'C9A84C'
WHITE  = 'FFFFFF'
LIGHT  = 'F5F0E8'
GREEN  = '1A7A4A'
RED    = 'C0392B'
ORANGE = 'E67E22'
GREY   = 'ECE9E0'

def style_header(cell, bg=NAVY, fg=WHITE, bold=True, size=11):
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=size, name='Calibri')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_cell(cell, bg=WHITE, fg='000000', bold=False, align='left'):
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=10, name='Calibri')
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ─────────────────────────────────────────────
# FEUILLE 1 — RÉSUMÉ EXÉCUTIF
# ─────────────────────────────────────────────
ws1 = wb.create_sheet('Résumé Exécutif')
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 45

# Titre
ws1.merge_cells('A1:B1')
ws1['A1'] = 'IMPACT ÉCONOMIQUE — SYSTÈME HYBRIDE BCOINS PLAY'
style_header(ws1['A1'], bg=NAVY, fg=GOLD, size=14)
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:B2')
ws1['A2'] = 'Analyse financière BarCoins — Avril 2026'
style_header(ws1['A2'], bg=GOLD, fg=NAVY, size=11)
ws1.row_dimensions[2].height = 22

# Verdict
ws1.merge_cells('A3:B3')
ws1['A3'] = '✅ VERDICT GLOBAL : GO — Lancer la beta juillet 2026'
style_header(ws1['A3'], bg=GREEN, fg=WHITE, size=12)
ws1.row_dimensions[3].height = 25

# KPIs clés
ws1.row_dimensions[4].height = 5
kpis = [
    ('Indicateur', 'Valeur'),
    ('MRR cible M12 (scenario central)', '6 800 €/mois'),
    ('ARR cible M12', '81 600 €'),
    ('LTV par bar', '3 913 €'),
    ('CAC estimé', '250 €'),
    ('Ratio LTV/CAC', '15,7x'),
    ('Churn mensuel CHR estimé', '2,3 %/mois'),
    ('Break-even (scenario central)', 'Mois 6'),
    ('Valorisation pré-money cible', '2,5 – 3,3 M€'),
    ('Besoin seed', '300 – 500 k€'),
    ('Dilution proposée', '10 – 15 %'),
    ('Bars actifs cibles M12 (central)', '50 bars'),
]
for i, (label, val) in enumerate(kpis):
    r = 5 + i
    ws1[f'A{r}'] = label
    ws1[f'B{r}'] = val
    if i == 0:
        style_header(ws1[f'A{r}'], bg=NAVY)
        style_header(ws1[f'B{r}'], bg=NAVY)
    else:
        bg = LIGHT if i % 2 == 0 else WHITE
        style_cell(ws1[f'A{r}'], bg=bg, bold=True)
        style_cell(ws1[f'B{r}'], bg=bg, align='center')
    ws1.row_dimensions[r].height = 20

apply_border(ws1, 5, 5 + len(kpis) - 1, 1, 2)

# Plans tarifaires
r = 5 + len(kpis) + 1
ws1.merge_cells(f'A{r}:B{r}')
ws1[f'A{r}'] = 'Plans tarifaires'
style_header(ws1[f'A{r}'], bg=NAVY)
ws1.row_dimensions[r].height = 22

plans = [
    ('STARTER', '89 €/mois'),
    ('STANDARD', '149 €/mois'),
    ('PREMIUM', '249 €/mois'),
]
for j, (plan, prix) in enumerate(plans):
    rr = r + 1 + j
    ws1[f'A{rr}'] = plan
    ws1[f'B{rr}'] = prix
    style_cell(ws1[f'A{rr}'], bg=LIGHT, bold=True)
    style_cell(ws1[f'B{rr}'], bg=LIGHT, align='center')
    ws1.row_dimensions[rr].height = 20

apply_border(ws1, r, r + len(plans), 1, 2)

# ─────────────────────────────────────────────
# FEUILLE 2 — BENCHMARKS PRIX
# ─────────────────────────────────────────────
ws2 = wb.create_sheet('Benchmarks Prix')
ws2.sheet_view.showGridLines = False
cols = [28, 15, 15, 16, 16, 16, 14]
for i, w in enumerate(cols):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

ws2.merge_cells('A1:G1')
ws2['A1'] = 'BENCHMARKS PRIX — Perpignan / Narbonne / Montpellier'
style_header(ws2['A1'], bg=NAVY, fg=GOLD, size=12)
ws2.row_dimensions[1].height = 30

headers = ['Produit', 'Perpignan', 'Narbonne', 'Montpellier', 'Source', 'Marge brute bar', 'Seuil Comptoir']
for i, h in enumerate(headers):
    c = ws2.cell(2, i+1, h)
    style_header(c, bg=GOLD, fg=NAVY)
ws2.row_dimensions[2].height = 25

data = [
    ('Shot',           '3,50 € [Hyp.]', '4,00 € [Sourcé]', '4,50 € [Sourcé]', 'TripAdvisor / menus publics', '75 %', '35 coins'),
    ('Bière 25cl',     '3,80 € [Sourcé]','4,20 € [Sourcé]', '5,00 € [Sourcé]', "O'Flaherty, menus 2026",     '77 %', '80 coins'),
    ('Bière qualité',  '5,50 € [Sourcé]','6,00 € [Hyp.]',  '6,50 € [Sourcé]', 'CraftBar Montpellier',        '68 %', '80 coins'),
    ('Cocktail',       '7,50 € [Sourcé]','8,00 € [Hyp.]',  '9,00 € [Sourcé]', 'Le Paradoxe, TripAdvisor',    '72 %', '160 coins'),
    ('Planche apéro',  '15,00 € [Sourcé]','17,00 € [Hyp.]','19,00 € [Sourcé]','Le Comptoir Perpignan',        '62 %', '200 coins'),
]
for i, row in enumerate(data):
    r = 3 + i
    for j, val in enumerate(row):
        c = ws2.cell(r, j+1, val)
        bg = LIGHT if i % 2 == 0 else WHITE
        style_cell(c, bg=bg, align='center' if j > 0 else 'left')
    ws2.row_dimensions[r].height = 22

apply_border(ws2, 2, 2 + len(data), 1, 7)

# Analyse cohérence
r2 = 3 + len(data) + 2
ws2.merge_cells(f'A{r2}:G{r2}')
ws2[f'A{r2}'] = 'Analyse cohérence seuils Comptoir'
style_header(ws2[f'A{r2}'], bg=NAVY)
ws2.row_dimensions[r2].height = 22

heads2 = ['Produit', 'Prix moyen', 'Seuil (coins)', 'Coins/visite (20€)', 'Visites pour seuil', 'Cohérence', '']
for i, h in enumerate(heads2[:6]):
    c = ws2.cell(r2+1, i+1, h)
    style_header(c, bg=GOLD, fg=NAVY)
ws2.row_dimensions[r2+1].height = 22

coh_data = [
    ('Shot',         '4,00 €', '35',  '4',  '9 visites',  '✅ Adapté'),
    ('Bière',        '4,33 €', '80',  '4',  '20 visites', '⚠️ Exigeant'),
    ('Cocktail',     '8,17 €', '160', '8',  '20 visites', '⚠️ Exigeant'),
    ('Planche apéro','17,00 €','200', '26', '8 visites',  '✅ Atteignable'),
]
for i, row in enumerate(coh_data):
    rr = r2 + 2 + i
    for j, val in enumerate(row):
        c = ws2.cell(rr, j+1, val)
        bg = LIGHT if i % 2 == 0 else WHITE
        fg = '000000'
        if '✅' in val:
            fg = GREEN
        elif '⚠️' in val:
            fg = ORANGE
        style_cell(c, bg=bg, align='center' if j > 0 else 'left', fg=fg)
    ws2.row_dimensions[rr].height = 22

apply_border(ws2, r2+1, r2+1+len(coh_data), 1, 6)

# ─────────────────────────────────────────────
# FEUILLE 3 — SCENARIOS FINANCIERS
# ─────────────────────────────────────────────
ws3 = wb.create_sheet('Scénarios Financiers')
ws3.sheet_view.showGridLines = False
scenario_cols = [14, 8, 10, 10, 10, 8, 10, 10, 14]
for i, w in enumerate(scenario_cols):
    ws3.column_dimensions[get_column_letter(i+1)].width = w

ws3.merge_cells('A1:I1')
ws3['A1'] = 'SCÉNARIOS FINANCIERS — Projections M1/M6/M12/M24 post-lancement (oct 2026)'
style_header(ws3['A1'], bg=NAVY, fg=GOLD, size=12)
ws3.row_dimensions[1].height = 30

# Hypothèses
ws3.merge_cells('A2:I2')
ws3['A2'] = 'Hypothèses : Mix plans STARTER 60% (89€) / STANDARD 30% (149€) / PREMIUM 10% (249€) — MRR moyen/bar = 116€ — CAC 250€ — Source churn CHR : Gira Conseil 2022'
style_header(ws3['A2'], bg=GREY, fg=NAVY, size=9, bold=False)
ws3.row_dimensions[2].height = 18

heads3 = ['Scénario', 'Période', 'Bars actifs', 'MRR (€)', 'ARR (€)', 'Churn/mois', 'LTV (€)', 'LTV/CAC', 'Break-even']
for i, h in enumerate(heads3):
    c = ws3.cell(3, i+1, h)
    style_header(c, bg=GOLD, fg=NAVY)
ws3.row_dimensions[3].height = 25

scenarios = [
    # (scenario, periode, bars, MRR, ARR, churn, LTV, ltv_cac, be)
    ('Prudent',   'M1',  5,    680,    8160,    '2,5%', 3264, '13,1x', '>24 mois'),
    ('Prudent',   'M6',  15,   2040,   24480,   '2,5%', 3264, '13,1x', '>24 mois'),
    ('Prudent',   'M12', 28,   3808,   45696,   '2,5%', 3264, '13,1x', 'Mois 13'),
    ('Prudent',   'M24', 55,   7484,   89808,   '2,5%', 3264, '13,1x', '—'),
    ('Central',   'M1',  8,    1088,   13056,   '2,3%', 3913, '15,7x', '>24 mois'),
    ('Central',   'M6',  25,   3400,   40800,   '2,3%', 3913, '15,7x', 'Mois 11'),
    ('Central',   'M12', 50,   6800,   81600,   '2,3%', 3913, '15,7x', 'Mois 6'),
    ('Central',   'M24', 120,  16320,  195840,  '2,3%', 3913, '15,7x', '—'),
    ('Ambitieux', 'M1',  12,   1632,   19584,   '2,0%', 4570, '18,3x', '>18 mois'),
    ('Ambitieux', 'M6',  40,   5440,   65280,   '2,0%', 4570, '18,3x', 'Mois 6'),
    ('Ambitieux', 'M12', 80,   10880,  130560,  '2,0%', 4570, '18,3x', 'Mois 4'),
    ('Ambitieux', 'M24', 200,  27200,  326400,  '2,0%', 4570, '18,3x', '—'),
]

scenario_colors = {'Prudent': 'DDEEFF', 'Central': 'DDFFEE', 'Ambitieux': 'FFF3DD'}
for i, row in enumerate(scenarios):
    r = 4 + i
    scenario = row[0]
    bg = scenario_colors.get(scenario, WHITE)
    for j, val in enumerate(row):
        c = ws3.cell(r, j+1, val)
        bold = j == 0
        align = 'center' if j > 0 else 'left'
        style_cell(c, bg=bg, bold=bold, align=align)
        if isinstance(val, int):
            c.number_format = '#,##0'
    ws3.row_dimensions[r].height = 20

apply_border(ws3, 3, 3 + len(scenarios), 1, 9)

# Légende
lr = 4 + len(scenarios) + 1
ws3.merge_cells(f'A{lr}:I{lr}')
ws3[f'A{lr}'] = '✅ VERDICT : GO — Ratio LTV/CAC >13x sur tous scénarios. Rentabilité entre M6 (ambitieux) et M13 (prudent). Seuil clé : 25 bars actifs.'
ws3[f'A{lr}'].fill = PatternFill('solid', fgColor=GREEN)
ws3[f'A{lr}'].font = Font(color=WHITE, bold=True, size=10, name='Calibri')
ws3[f'A{lr}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws3.row_dimensions[lr].height = 25

# ─────────────────────────────────────────────
# FEUILLE 4 — MODÈLE UNITAIRE BCOINS
# ─────────────────────────────────────────────
ws4 = wb.create_sheet('Modèle Unitaire Bcoins')
ws4.sheet_view.showGridLines = False
for i, w in enumerate([22, 18, 14, 14, 14, 14]):
    ws4.column_dimensions[get_column_letter(i+1)].width = w

ws4.merge_cells('A1:F1')
ws4['A1'] = 'MODÈLE UNITAIRE — Coins, Rangs, Coût Réel, Dette Promotionnelle'
style_header(ws4['A1'], bg=NAVY, fg=GOLD, size=12)
ws4.row_dimensions[1].height = 30

# Conversion montant -> coins
ws4.merge_cells('A2:F2')
ws4['A2'] = 'Conversion Montant → playCoins'
style_header(ws4['A2'], bg=NAVY)
ws4.row_dimensions[2].height = 22

heads4 = ['Addition (€)', 'Palier', 'Multiplicateur', 'playCoins octroyés', '', '']
for i, h in enumerate(heads4[:4]):
    c = ws4.cell(3, i+1, h)
    style_header(c, bg=GOLD, fg=NAVY)
ws4.row_dimensions[3].height = 22

conv = [
    ('10 €',  '10–19,99 €', '×1,5', '15'),
    ('20 €',  '20–29,99 €', '×2',   '40'),
    ('35 €',  '30–49,99 €', '×3',   '105'),
    ('50 €',  '50 €+',      '×4',   '200'),
]
for i, row in enumerate(conv):
    r = 4 + i
    for j, val in enumerate(row):
        c = ws4.cell(r, j+1, val)
        bg = LIGHT if i % 2 == 0 else WHITE
        style_cell(c, bg=bg, align='center')
    ws4.row_dimensions[r].height = 20
apply_border(ws4, 3, 3+len(conv), 1, 4)

# Rangs Aven-Jers
r5 = 4 + len(conv) + 2
ws4.merge_cells(f'A{r5}:F{r5}')
ws4[f'A{r5}'] = 'Rangs Aven-Jers — Visites nécessaires par profil'
style_header(ws4[f'A{r5}'], bg=NAVY)
ws4.row_dimensions[r5].height = 22

heads5 = ['Rang', 'Seuil coins', 'Note 10€', 'Note 20€', 'Note 35€', 'Note 50€']
for i, h in enumerate(heads5):
    c = ws4.cell(r5+1, i+1, h)
    style_header(c, bg=GOLD, fg=NAVY)
ws4.row_dimensions[r5+1].height = 22

rangs = [
    ('Captain Americano N1', '30',   '2',   '1',  '1', '1'),
    ('Ant-Rhum N2',          '200',  '14',  '5',  '2', '1'),
    ('Lo-Kir N3',            '600',  '40',  '15', '6', '3'),
    ('Spider-Malt N4',       '1500', '100', '38', '15','8'),
    ('Thorquila N5',         '3500', '234', '88', '34','18'),
    ('Iron Martini N6',      '7000', '467', '175','67','35'),
    ('Nick Fu-Rhum N7',      '12000','800', '300','115','60'),
]
for i, row in enumerate(rangs):
    rr = r5 + 2 + i
    for j, val in enumerate(row):
        c = ws4.cell(rr, j+1, val)
        bg = LIGHT if i % 2 == 0 else WHITE
        style_cell(c, bg=bg, align='center' if j > 0 else 'left')
    ws4.row_dimensions[rr].height = 20
apply_border(ws4, r5+1, r5+1+len(rangs), 1, 6)

# Dette promo
rd = r5 + 2 + len(rangs) + 2
ws4.merge_cells(f'A{rd}:F{rd}')
ws4[f'A{rd}'] = 'Dette Promotionnelle — 100 joueurs actifs/mois'
style_header(ws4[f'A{rd}'], bg=NAVY)
ws4.row_dimensions[rd].height = 22

dette = [
    ('Ticket moyen', '20 €'),
    ('Visites/mois/joueur', '3'),
    ('Coins émis/mois (100 joueurs)', '6 000'),
    ('Redemptions max (shots à 35c)', '171 shots offerts'),
    ('Coût réel max pour le bar', '171 € / mois'),
    ('Verdict', '✅ <10% du CA additionnel généré'),
]
for i, (label, val) in enumerate(dette):
    rr = rd + 1 + i
    ws4[f'A{rr}'] = label
    ws4[f'B{rr}'] = val
    ws4.merge_cells(f'B{rr}:F{rr}')
    bg = LIGHT if i % 2 == 0 else WHITE
    fg = GREEN if '✅' in val else '000000'
    style_cell(ws4[f'A{rr}'], bg=bg, bold=True)
    style_cell(ws4[f'B{rr}'], bg=bg, align='center', fg=fg)
    ws4.row_dimensions[rr].height = 20
apply_border(ws4, rd+1, rd+len(dette), 1, 6)

# ─────────────────────────────────────────────
# FEUILLE 5 — RISQUES
# ─────────────────────────────────────────────
ws5 = wb.create_sheet('Risques')
ws5.sheet_view.showGridLines = False
for i, w in enumerate([28, 14, 18, 42]):
    ws5.column_dimensions[get_column_letter(i+1)].width = w

ws5.merge_cells('A1:D1')
ws5['A1'] = 'RISQUES FINANCIERS — Top 5 + Mitigations'
style_header(ws5['A1'], bg=NAVY, fg=GOLD, size=12)
ws5.row_dimensions[1].height = 30

heads_r = ['Risque', 'Probabilité', 'Impact (€)', 'Mitigation']
for i, h in enumerate(heads_r):
    c = ws5.cell(2, i+1, h)
    style_header(c, bg=GOLD, fg=NAVY)
ws5.row_dimensions[2].height = 22

risks = [
    ('Dette promotionnelle excessive', 'Moyenne', '6 000 €/bar/an', 'Plafond coinsTotal, expiration annuelle, alertes ratio redemption'),
    ('Concentration 5 bars pilotes', 'Haute', '35–50 % MRR', 'Diversification rapide : objectif 15 bars à M+3, pacte stabilité pilote'),
    ('Rejet validation ANJ', 'Faible', '−35 à −40 % CA potentiel', 'Architecture modulaire, pivot Comptoir/Fidélité sans ANJ'),
    ('Churn élevé gérants CHR', 'Moyenne', 'CAC × résiliés (250€/bar)', 'Scoring churn, offres rétention, programme ambassadeur'),
    ('Délai adoption client final', 'Moyenne', '−25 % croissance', 'Missions "booster", events lancement, influenceurs régionaux'),
]
prob_colors = {'Haute': RED, 'Moyenne': ORANGE, 'Faible': GREEN}
for i, row in enumerate(risks):
    r = 3 + i
    ws5.cell(r, 1, row[0])
    style_cell(ws5.cell(r, 1), bg=LIGHT if i%2==0 else WHITE, bold=True)
    ws5.cell(r, 2, row[1])
    style_cell(ws5.cell(r, 2), bg=prob_colors.get(row[1], WHITE), fg=WHITE, align='center', bold=True)
    ws5.cell(r, 3, row[2])
    style_cell(ws5.cell(r, 3), bg=LIGHT if i%2==0 else WHITE, align='center')
    ws5.cell(r, 4, row[3])
    style_cell(ws5.cell(r, 4), bg=LIGHT if i%2==0 else WHITE)
    ws5.row_dimensions[r].height = 40
apply_border(ws5, 2, 2+len(risks), 1, 4)

# ─────────────────────────────────────────────
# FEUILLE 6 — VALORISATION & NEXT STEPS
# ─────────────────────────────────────────────
ws6 = wb.create_sheet('Valorisation & Next Steps')
ws6.sheet_view.showGridLines = False
for i, w in enumerate([30, 35]):
    ws6.column_dimensions[get_column_letter(i+1)].width = w

ws6.merge_cells('A1:B1')
ws6['A1'] = 'VALORISATION & RECOMMANDATIONS INVESTISSEUR'
style_header(ws6['A1'], bg=NAVY, fg=GOLD, size=12)
ws6.row_dimensions[1].height = 30

valo = [
    ('Méthode', 'Valorisation estimée'),
    ('Multiple ARR SaaS (×5 à ×8)', '275 k – 1 M€'),
    ('DCF simplifié 3 ans', '350 k – 450 k€'),
    ('Benchmark startups CHR', '250 k – 640 k€'),
    ('Cible retenue (post-proof 50 bars)', '2,5 – 3,3 M€ pré-money'),
]
for i, (label, val) in enumerate(valo):
    r = 2 + i
    ws6[f'A{r}'] = label
    ws6[f'B{r}'] = val
    if i == 0:
        style_header(ws6[f'A{r}'], bg=GOLD, fg=NAVY)
        style_header(ws6[f'B{r}'], bg=GOLD, fg=NAVY)
    else:
        bg = LIGHT if i%2==0 else WHITE
        bold = i == len(valo)-1
        style_cell(ws6[f'A{r}'], bg=bg, bold=bold)
        style_cell(ws6[f'B{r}'], bg=bg, bold=bold, align='center')
    ws6.row_dimensions[r].height = 22
apply_border(ws6, 2, 2+len(valo)-1, 1, 2)

# Financement
rf = 2 + len(valo) + 1
ws6.merge_cells(f'A{rf}:B{rf}')
ws6[f'A{rf}'] = 'Structure de financement'
style_header(ws6[f'A{rf}'], bg=NAVY)
ws6.row_dimensions[rf].height = 22

fonds = [
    ('Phase 1 — Beta (maintenant)', '12 – 15 k€ — Juridique, ANJ, intégrations'),
    ('Phase 2 — Seed (post-50 bars)', '300 – 500 k€ — Prod 40% / Sales 30% / Legal 10% / Ops 10% / BFR 10%'),
    ('Dilution proposée', '10 – 15 % post-money'),
]
for i, (label, val) in enumerate(fonds):
    rr = rf + 1 + i
    ws6[f'A{rr}'] = label
    ws6[f'B{rr}'] = val
    bg = LIGHT if i%2==0 else WHITE
    style_cell(ws6[f'A{rr}'], bg=bg, bold=True)
    style_cell(ws6[f'B{rr}'], bg=bg)
    ws6.row_dimensions[rr].height = 30
apply_border(ws6, rf+1, rf+len(fonds), 1, 2)

# Scoring
rs = rf + len(fonds) + 2
ws6.merge_cells(f'A{rs}:B{rs}')
ws6[f'A{rs}'] = 'Scoring GO/NO-GO Investisseur'
style_header(ws6[f'A{rs}'], bg=NAVY)
ws6.row_dimensions[rs].height = 22

scores = [
    ('Critère', 'Score'),
    ('Marché', '5/5 — >15k bars France, niche sous-adressée'),
    ('Exécution', '4/5 — Solo founder, roadmap claire, partenaires pilotes'),
    ('Timing', '4/5 — Post-Covid, digitalisation CHR accélérée'),
]
for i, (label, val) in enumerate(scores):
    rr = rs + 1 + i
    ws6[f'A{rr}'] = label
    ws6[f'B{rr}'] = val
    if i == 0:
        style_header(ws6[f'A{rr}'], bg=GOLD, fg=NAVY)
        style_header(ws6[f'B{rr}'], bg=GOLD, fg=NAVY)
    else:
        bg = LIGHT if i%2==0 else WHITE
        style_cell(ws6[f'A{rr}'], bg=bg, bold=True)
        style_cell(ws6[f'B{rr}'], bg=bg)
    ws6.row_dimensions[rr].height = 22
apply_border(ws6, rs+1, rs+len(scores), 1, 2)

# Next steps
rn = rs + len(scores) + 2
ws6.merge_cells(f'A{rn}:B{rn}')
ws6[f'A{rn}'] = 'Next Steps — Cette semaine'
style_header(ws6[f'A{rn}'], bg=NAVY)
ws6.row_dimensions[rn].height = 22

steps = [
    ('1. Finaliser signature 5 bars pilotes', '+ démo live investisseurs'),
    ('2. DCF détaillé & data room', 'Chiffres terrain propres'),
    ('3. Lettre intention ANJ', 'Roadmap SaaS hors-jeu prête'),
    ('4. Premier RDV BPI / angel local', 'Booker cette semaine'),
    ('5. Kit pre-seed', 'Deck + modèle financier + term sheet'),
]
for i, (label, val) in enumerate(steps):
    rr = rn + 1 + i
    ws6[f'A{rr}'] = label
    ws6[f'B{rr}'] = val
    bg = LIGHT if i%2==0 else WHITE
    style_cell(ws6[f'A{rr}'], bg=bg, bold=True)
    style_cell(ws6[f'B{rr}'], bg=bg)
    ws6.row_dimensions[rr].height = 22
apply_border(ws6, rn+1, rn+len(steps), 1, 2)

# Verdict final
rv = rn + len(steps) + 2
ws6.merge_cells(f'A{rv}:B{rv}')
ws6[f'A{rv}'] = '✅ VERDICT FINAL : GO ABSOLU — Lancer beta juillet 2026, capper dilution, prouver repeat revenue sur 30 bars actifs'
ws6[f'A{rv}'].fill = PatternFill('solid', fgColor=GREEN)
ws6[f'A{rv}'].font = Font(color=WHITE, bold=True, size=11, name='Calibri')
ws6[f'A{rv}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws6.row_dimensions[rv].height = 35

wb.save(OUTPUT)
print('Excel genere : ' + OUTPUT)
print('Taille : ' + str(os.path.getsize(OUTPUT)) + ' bytes')
